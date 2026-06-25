"""Generate exhaustive decision-equivalent reports from the current rule engine.

The integer inputs are unbounded, so the report uses representative values for
each distinct branch of the implemented logic:

- n_overlaid_series: 1 (no warning), 4 (readability warning), 7 (split/filter)
- n_comparison_levels: 2 (paired dot/slope), 3 (>2 repeated-measures line)
- n_compositional_parts: 2, 3 (ternary eligible), 4 (>3 parts)

All categorical and boolean values are enumerated exhaustively.
"""

import csv
import json
from collections import Counter
from dataclasses import asdict, fields
from itertools import product
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from decision_tree import (
    ALLOWED_VALUES,
    DecisionInputs,
    Recommendation,
    REPORT_REPRESENTATIVE_VALUES,
    format_result,
    recommend_visualisations,
)


ROOT = Path(__file__).resolve().parent
REPORT_DIR = ROOT / "decision_report"

VALID_CSV = REPORT_DIR / "valid_combinations.csv"
LONG_CSV = REPORT_DIR / "recommendations_long.csv"
RECOMMENDATION_SET_CSV = REPORT_DIR / "recommendation_sets.csv"
VALIDATION_CSV = REPORT_DIR / "validation_examples.csv"
SUMMARY_JSON = REPORT_DIR / "report_summary.json"
WORKBOOK_JSON = REPORT_DIR / "workbook_data.json"
METHODOLOGY_MD = ROOT / "Decision_tree_full_report.md"
WORKBOOK_XLSX = ROOT / "Decision_tree_full_report.xlsx"

INPUT_FIELDS = [field.name for field in fields(DecisionInputs)]
RECOMMENDATION_FIELDS = [field.name for field in fields(Recommendation)]

VALID_HEADERS = [
    "combination_id",
    *INPUT_FIELDS,
    "decision_path",
    "recommendation_count",
    "recommendation_names",
    "recommendation_ranks",
    "recommendation_output",
    "design_notes",
]

LONG_HEADERS = [
    "combination_id",
    *INPUT_FIELDS,
    "recommendation_number",
    *RECOMMENDATION_FIELDS,
    "design_notes",
]

RECOMMENDATION_SET_HEADERS = [
    "recommendation_set_id",
    "valid_combination_count",
    "primary_tasks",
    "data_forms",
    "target_audiences",
    "display_levels",
    "comparison_foci",
    "comparison_structures",
    "temporal_contexts",
    "n_overlaid_series_values",
    "n_comparison_levels_values",
    "n_compositional_parts_values",
    "recommendation_count",
    "recommendation_names",
    "recommendation_visual_mappings",
    "recommendation_ranks",
    "recommendation_rationales",
    "recommendation_use_when",
    "recommendation_cautions",
    "recommendation_example_images",
    "recommendation_example_sources",
    "recommendation_example_code_files",
    "representative_combination_id",
    "representative_decision_path",
    *[f"example_{field}" for field in INPUT_FIELDS],
    "review_status",
    "review_notes",
]


def enumeration_domains():
    """Discover the reporting domain for every DecisionInputs field.

    Adding or removing a supported field in ``DecisionInputs`` automatically
    changes report enumeration. New categorical fields must be added to
    ``ALLOWED_VALUES``; new boolean fields need no extra metadata; numeric or
    otherwise unbounded fields need representative values in
    ``REPORT_REPRESENTATIVE_VALUES``. Unsupported fields fail loudly.
    """

    domains = {}
    for field in fields(DecisionInputs):
        if field.name in ALLOWED_VALUES:
            domains[field.name] = sorted(ALLOWED_VALUES[field.name])
        elif field.type is bool:
            domains[field.name] = [False, True]
        elif field.name in REPORT_REPRESENTATIVE_VALUES:
            domains[field.name] = list(REPORT_REPRESENTATIVE_VALUES[field.name])
        else:
            raise RuntimeError(
                f"Report enumeration does not know how to enumerate DecisionInputs "
                f"field '{field.name}'. Add accepted categorical values to "
                "ALLOWED_VALUES, use a boolean field, or add decision-equivalent "
                "representative values to REPORT_REPRESENTATIVE_VALUES in "
                "decision_tree.py."
            )
    return domains


def parameter_report_rows():
    """Describe the dynamically discovered report domains."""

    domains = enumeration_domains()
    rows = []
    for field in fields(DecisionInputs):
        values = domains[field.name]
        if field.name in ALLOWED_VALUES:
            enumeration_type = "Exhaustive categorical"
            interpretation = "All accepted values from decision_tree.py."
        elif field.type is bool:
            enumeration_type = "Exhaustive boolean"
            interpretation = "Both boolean values."
        else:
            enumeration_type = "Decision-equivalent representatives"
            interpretation = " | ".join(
                str(description)
                for description in REPORT_REPRESENTATIVE_VALUES[field.name].values()
            )
        rows.append(
            {
                "parameter": field.name,
                "values_enumerated": "; ".join(str(value) for value in values),
                "enumeration_type": enumeration_type,
                "interpretation": interpretation,
            }
        )
    return rows


def candidate_inputs():
    """Yield all enumerated combinations; validation identifies compatible paths."""

    domains = enumeration_domains()
    field_names = list(domains)
    for values in product(*(domains[name] for name in field_names)):
        yield DecisionInputs(**dict(zip(field_names, values)))


def validation_examples():
    """Return the first enumerated example for every distinct validation message."""

    examples_by_message = {}
    for inputs in candidate_inputs():
        try:
            recommend_visualisations(inputs)
        except (TypeError, ValueError) as error:
            message = str(error)
            examples_by_message.setdefault(
                message,
                {
                    "example": f"Representative rejected combination {len(examples_by_message) + 1}",
                    **asdict(inputs),
                    "corrective_error_message": message,
                },
            )
    return list(examples_by_message.values())


def write_csv(path, headers, rows):
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=headers, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(rows)


def recommendation_signature(recommendations):
    """Return the review-level identity of a recommendation output."""

    return tuple(
        json.dumps(asdict(rec), sort_keys=True)
        for rec in recommendations
    )


def sorted_join(values):
    """Format a set of values consistently for reports."""

    return " | ".join(
        "None" if value is None else str(value)
        for value in sorted(values, key=lambda item: (item is not None, str(item)))
    )


def recommendation_set_rows(recommendation_sets):
    """Return one row for each distinct ordered recommendation detail set."""

    rows = []
    for index, record in enumerate(
        sorted(
            recommendation_sets.values(),
            key=lambda item: (-item["valid_combination_count"], item["recommendation_names"]),
        ),
        start=1,
    ):
        example_inputs = {
            f"example_{field}": record["example_inputs"][field]
            for field in INPUT_FIELDS
        }
        rows.append(
            {
                "recommendation_set_id": f"RS{index:03d}",
                "valid_combination_count": record["valid_combination_count"],
                "primary_tasks": sorted_join(record["primary_tasks"]),
                "data_forms": sorted_join(record["data_forms"]),
                "target_audiences": sorted_join(record["target_audiences"]),
                "display_levels": sorted_join(record["display_levels"]),
                "comparison_foci": sorted_join(record["comparison_foci"]),
                "comparison_structures": sorted_join(record["comparison_structures"]),
                "temporal_contexts": sorted_join(record["temporal_contexts"]),
                "n_overlaid_series_values": sorted_join(record["n_overlaid_series_values"]),
                "n_comparison_levels_values": sorted_join(record["n_comparison_levels_values"]),
                "n_compositional_parts_values": sorted_join(record["n_compositional_parts_values"]),
                "recommendation_count": len(record["recommendation_names"]),
                "recommendation_names": " | ".join(record["recommendation_names"]),
                "recommendation_visual_mappings": " | ".join(
                    record["recommendation_visual_mappings"]
                ),
                "recommendation_ranks": " | ".join(record["recommendation_ranks"]),
                "recommendation_rationales": " | ".join(record["recommendation_rationales"]),
                "recommendation_use_when": " | ".join(record["recommendation_use_when"]),
                "recommendation_cautions": " | ".join(record["recommendation_cautions"]),
                "recommendation_example_images": " | ".join(
                    record["recommendation_example_images"]
                ),
                "recommendation_example_sources": " | ".join(
                    record["recommendation_example_sources"]
                ),
                "recommendation_example_code_files": " | ".join(
                    record["recommendation_example_code_files"]
                ),
                "representative_combination_id": record["representative_combination_id"],
                "representative_decision_path": record["representative_decision_path"],
                **example_inputs,
                "review_status": "To review",
                "review_notes": "",
            }
        )
    return rows


def summary_metric_rows(summary):
    """Return the report metrics shown in both the workbook and Markdown."""

    metrics = [
        ("Candidate combinations evaluated", "candidate_combinations_evaluated"),
        (
            "Valid decision-equivalent combinations",
            "valid_decision_equivalent_combinations",
        ),
        ("Rejected candidate combinations", "rejected_candidate_combinations"),
        ("Long-format recommendation rows", "long_format_recommendation_rows"),
        ("Distinct chart names", "unique_visualisation_names"),
        (
            "Distinct chart-name combinations",
            "unique_ordered_recommendation_name_sets",
        ),
        (
            "Distinct full recommendation outputs",
            "unique_reviewable_recommendation_sets",
        ),
        (
            "Distinct individual recommendation cards",
            "unique_recommendation_detail_records",
        ),
        (
            "Unique recommendation-plus-guidance outputs",
            "unique_recommendation_and_guidance_outputs",
        ),
        ("Unique complete formatted outputs", "unique_complete_formatted_outputs"),
        ("Unique decision paths", "unique_decision_paths"),
        ("Unique design-note sets", "unique_design_note_sets"),
    ]
    return [
        {"report_metric": label, "count": summary[key]}
        for label, key in metrics
    ]


def how_to_use_rows():
    """Return the shared report navigation notes."""

    return [
        {
            "section": "Valid Combinations",
            "description": (
                "One row per valid parameter combination, including the complete "
                "formatted output."
            ),
        },
        {
            "section": "Recommendations Long",
            "description": (
                "One row per individual recommendation; best for filtering, "
                "counting, and rule auditing."
            ),
        },
        {
            "section": "Recommendation Sets",
            "description": (
                "One row per distinct full recommendation output, including "
                "chart names, ranks, visual mappings, rationales, cautions, "
                "and use-when statements; best for manual review of end-point "
                "adequacy."
            ),
        },
        {
            "section": "Parameters",
            "description": (
                "Documents accepted values and representative numeric equivalence "
                "classes."
            ),
        },
        {
            "section": "Validation Examples",
            "description": (
                "Representative user mistakes and the corrective messages produced "
                "by the code."
            ),
        },
        {
            "section": "Rejected Reasons",
            "description": (
                "Counts of rejected candidate combinations by validation reason."
            ),
        },
        {
            "section": "Reproducibility",
            "description": (
                "Regenerate using generate_decision_report.py after changing "
                "decision_tree.py."
            ),
        },
        {
            "section": "Uniqueness counts",
            "description": (
                "Distinct chart names count labels only. Chart-name combinations "
                "count ordered bundles of labels. Full recommendation outputs "
                "also include details such as rank, mapping, rationale, use-when "
                "text, cautions, and implementation guidance."
            ),
        },
    ]


def markdown_escape(value):
    if value is None:
        return "None"
    return str(value).replace("|", "\\|").replace("\n", "<br>")


def markdown_table(headers, rows):
    lines = [
        "| " + " | ".join(headers) + " |",
        "| " + " | ".join("---" for _ in headers) + " |",
    ]
    for row in rows:
        if isinstance(row, dict):
            values = [row.get(header, "") for header in headers]
        else:
            values = row
        lines.append(
            "| " + " | ".join(markdown_escape(value) for value in values) + " |"
        )
    return "\n".join(lines)


def _normalise_excel_value(value):
    """Return a scalar value suitable for writing to an Excel cell."""

    if isinstance(value, (list, tuple, set)):
        return " | ".join(str(item) for item in value)
    if isinstance(value, dict):
        return json.dumps(value, ensure_ascii=False, sort_keys=True)
    return value


def _write_excel_sheet(workbook, title, headers, rows):
    """Write one tabular sheet with simple filtering and readable widths."""

    sheet = workbook.create_sheet(title=title[:31])
    sheet.append(headers)
    header_fill = PatternFill("solid", fgColor="DCEFE9")
    for cell in sheet[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill

    for row in rows:
        sheet.append([_normalise_excel_value(row.get(header, "")) for header in headers])

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions

    for column_cells in sheet.columns:
        column_letter = get_column_letter(column_cells[0].column)
        max_length = max(
            len(str(cell.value)) if cell.value is not None else 0
            for cell in column_cells[:200]
        )
        sheet.column_dimensions[column_letter].width = min(max(max_length + 2, 12), 55)


def write_excel_workbook(workbook_data):
    """Write the companion Excel workbook from the same report data."""

    workbook = Workbook()
    workbook.remove(workbook.active)

    summary_rows = []
    summary_rows.append({"section": "Report metrics", "item": "", "value": ""})
    summary_rows.extend(
        {
            "section": "Report metrics",
            "item": row["report_metric"],
            "value": row["count"],
        }
        for row in workbook_data["summary_metric_rows"]
    )
    summary_rows.append({"section": "Valid combinations by primary task", "item": "", "value": ""})
    summary_rows.extend(
        {
            "section": "Valid combinations by primary task",
            "item": row["primary_task"],
            "value": row["valid_combinations"],
        }
        for row in workbook_data["task_rows"]
    )
    summary_rows.append({"section": "Recommendation frequency", "item": "", "value": ""})
    summary_rows.extend(
        {
            "section": "Recommendation frequency",
            "item": row["recommendation_name"],
            "value": row["frequency"],
        }
        for row in workbook_data["recommendation_frequency_rows"]
    )
    summary_rows.append({"section": "How to use this report", "item": "", "value": ""})
    summary_rows.extend(
        {
            "section": "How to use this report",
            "item": row["section"],
            "value": row["description"],
        }
        for row in workbook_data["how_to_use_rows"]
    )

    _write_excel_sheet(workbook, "Summary", ["section", "item", "value"], summary_rows)
    _write_excel_sheet(
        workbook,
        "Parameters",
        ["parameter", "values_enumerated", "enumeration_type", "interpretation"],
        workbook_data["parameter_rows"],
    )
    _write_excel_sheet(
        workbook,
        "Valid Combinations",
        workbook_data["valid_headers"],
        workbook_data["valid_rows"],
    )
    _write_excel_sheet(
        workbook,
        "Recommendations Long",
        workbook_data["long_headers"],
        workbook_data["long_rows"],
    )
    _write_excel_sheet(
        workbook,
        "Recommendation Sets",
        workbook_data["recommendation_set_headers"],
        workbook_data["recommendation_set_rows"],
    )
    _write_excel_sheet(
        workbook,
        "Validation Examples",
        workbook_data["validation_headers"],
        workbook_data["validation_rows"],
    )
    _write_excel_sheet(
        workbook,
        "Rejected Reasons",
        ["reason", "count"],
        workbook_data["rejected_reason_rows"],
    )

    workbook.save(WORKBOOK_XLSX)


def write_markdown_report(workbook_data):
    """Write a Markdown report that mirrors the workbook sections."""

    summary = workbook_data["summary"]
    metric_rows = [
        {
            "Report metric": row["report_metric"],
            "Count": f"{row['count']:,}",
        }
        for row in workbook_data["summary_metric_rows"]
    ]
    task_rows = [
        {
            "Primary task": row["primary_task"],
            "Valid combinations": f"{row['valid_combinations']:,}",
        }
        for row in workbook_data["task_rows"]
    ]
    recommendation_rows = [
        {
            "Recommendation name": row["recommendation_name"],
            "Frequency": f"{row['frequency']:,}",
        }
        for row in workbook_data["recommendation_frequency_rows"]
    ]
    report_navigation_rows = [
        {
            "Section": row["section"],
            "Description": row["description"],
        }
        for row in workbook_data["how_to_use_rows"]
    ]
    rejected_rows = [
        {
            "Reason": row["reason"],
            "Count": f"{row['count']:,}",
        }
        for row in workbook_data["rejected_reason_rows"]
    ]

    validation_rows = [
        {
            header: row.get(header)
            for header in workbook_data["validation_headers"]
        }
        for row in workbook_data["validation_rows"]
    ]

    valid_columns = ", ".join(f"`{header}`" for header in workbook_data["valid_headers"])
    long_columns = ", ".join(f"`{header}`" for header in workbook_data["long_headers"])
    recommendation_set_columns = ", ".join(
        f"`{header}`" for header in workbook_data["recommendation_set_headers"]
    )

    methodology = f"""# Decision support for visualising accelerometer-derived movement behaviour data

This Markdown report mirrors the companion Excel workbook:
[`Decision_tree_full_report.xlsx`](Decision_tree_full_report.xlsx).

This is the full decision-tree recommendation report for the decision-support
component.

The full large tables are kept as CSV files so the Markdown remains readable.
The workbook contains the same sections as separate sheets with filters and
formatting.

## Workbook-equivalent contents

{markdown_table(
    ["Excel sheet", "Markdown section", "Full data"],
    [
        ["Summary", "Summary", "Inline below"],
        ["Parameters", "Parameters", "Inline below"],
        [
            "Valid Combinations",
            "Valid Combinations",
            "[valid_combinations.csv](decision_report/valid_combinations.csv)",
        ],
        [
            "Recommendations Long",
            "Recommendations Long",
            "[recommendations_long.csv](decision_report/recommendations_long.csv)",
        ],
        [
            "Recommendation Sets",
            "Recommendation Sets",
            "[recommendation_sets.csv](decision_report/recommendation_sets.csv)",
        ],
        [
            "Validation Examples",
            "Validation Examples",
            "[validation_examples.csv](decision_report/validation_examples.csv)",
        ],
        ["Rejected Reasons", "Rejected Reasons", "Inline below"],
    ],
)}

## Purpose

This report documents the visualisation recommendations produced by the current
rule-based decision-tree implementation for all enumerated decision-equivalent
valid input combinations.

## Enumeration approach

All accepted categorical values and booleans were represented. Integer inputs
are theoretically unbounded, so representative values were selected for every
distinct implemented logic outcome. The report therefore enumerates every
distinct recommendation pathway produced by the implemented rules, rather than
every possible integer value.

## Summary

### Report metrics

{markdown_table(["Report metric", "Count"], metric_rows)}

### Valid combinations by primary task

{markdown_table(["Primary task", "Valid combinations"], task_rows)}

### Recommendation frequency

{markdown_table(["Recommendation name", "Frequency"], recommendation_rows)}

### How to use this report

{markdown_table(["Section", "Description"], report_navigation_rows)}

## Parameters

{markdown_table(
    ["parameter", "values_enumerated", "enumeration_type", "interpretation"],
    workbook_data["parameter_rows"],
)}

## Valid Combinations

Full table: [decision_report/valid_combinations.csv](decision_report/valid_combinations.csv)

- Rows: **{summary['valid_decision_equivalent_combinations']:,}**
- Columns: {valid_columns}

![Valid Combinations preview](decision_report/Valid_Combinations_preview.png)

## Recommendations Long

Full table: [decision_report/recommendations_long.csv](decision_report/recommendations_long.csv)

- Rows: **{summary['long_format_recommendation_rows']:,}**
- Columns: {long_columns}

![Recommendations Long preview](decision_report/Recommendations_Long_preview.png)

## Recommendation Sets

Full table: [decision_report/recommendation_sets.csv](decision_report/recommendation_sets.csv)

- Rows: **{summary['unique_reviewable_recommendation_sets']:,}**
- Columns: {recommendation_set_columns}

![Recommendation Sets preview](decision_report/Recommendation_Sets_preview.png)

## Validation Examples

Full table: [decision_report/validation_examples.csv](decision_report/validation_examples.csv)

{markdown_table(workbook_data["validation_headers"], validation_rows)}

## Rejected Reasons

{markdown_table(["Reason", "Count"], rejected_rows)}

## Reproducibility

The artifacts were generated from `decision_tree.py` using
`generate_decision_report.py`. Rerunning the generator after changing the rule
engine updates the workbook, Markdown report, CSV files, JSON files, and sheet
previews. The generator discovers the `DecisionInputs` fields, categorical
domains, booleans, and report-specific representative values from
`decision_tree.py`. If a new field cannot be enumerated, generation stops with
a corrective error rather than silently omitting it.
"""
    METHODOLOGY_MD.write_text(methodology, encoding="utf-8")


def main():
    valid_rows = []
    long_rows = []
    rejected_reasons = Counter()
    rejected_examples = {}
    recommendation_counts = Counter()
    task_counts = Counter()
    unique_ordered_recommendation_name_sets = set()
    unique_recommendation_detail_records = set()
    unique_recommendation_and_guidance_outputs = set()
    unique_complete_formatted_outputs = set()
    unique_decision_paths = set()
    unique_design_note_sets = set()
    recommendation_sets = {}
    candidate_count = 0

    for inputs in candidate_inputs():
        candidate_count += 1
        try:
            result = recommend_visualisations(inputs)
        except (TypeError, ValueError) as error:
            message = str(error)
            rejected_reasons[message] += 1
            rejected_examples.setdefault(
                message,
                {
                    "example": f"Representative rejected combination {len(rejected_examples) + 1}",
                    **asdict(inputs),
                    "corrective_error_message": message,
                },
            )
            continue

        combination_id = f"C{len(valid_rows) + 1:06d}"
        input_values = asdict(inputs)
        names = [rec.visualisation for rec in result.recommendations]
        ranks = [rec.rank for rec in result.recommendations]
        notes = " | ".join(result.design_notes)
        recommendation_details = tuple(
            json.dumps(asdict(rec), sort_keys=True) for rec in result.recommendations
        )
        review_signature = recommendation_signature(result.recommendations)
        design_note_signature = tuple(result.design_notes)
        unique_ordered_recommendation_name_sets.add(tuple(names))
        unique_recommendation_detail_records.update(recommendation_details)
        unique_recommendation_and_guidance_outputs.add(
            (recommendation_details, design_note_signature)
        )
        unique_complete_formatted_outputs.add(format_result(result))
        unique_decision_paths.add(tuple(result.decision_path))
        unique_design_note_sets.add(design_note_signature)
        recommendation_set = recommendation_sets.setdefault(
            review_signature,
            {
                "valid_combination_count": 0,
                "primary_tasks": set(),
                "data_forms": set(),
                "target_audiences": set(),
                "display_levels": set(),
                "comparison_foci": set(),
                "comparison_structures": set(),
                "temporal_contexts": set(),
                "n_overlaid_series_values": set(),
                "n_comparison_levels_values": set(),
                "n_compositional_parts_values": set(),
                "recommendation_names": names,
                "recommendation_visual_mappings": [
                    rec.visual_mapping for rec in result.recommendations
                ],
                "recommendation_ranks": ranks,
                "recommendation_rationales": [
                    rec.rationale for rec in result.recommendations
                ],
                "recommendation_use_when": [
                    rec.use_when for rec in result.recommendations
                ],
                "recommendation_cautions": [
                    rec.caution or "" for rec in result.recommendations
                ],
                "recommendation_example_images": [
                    rec.example_image_file or "" for rec in result.recommendations
                ],
                "recommendation_example_sources": [
                    rec.example_source or "" for rec in result.recommendations
                ],
                "recommendation_example_code_files": [
                    rec.example_code_file or "" for rec in result.recommendations
                ],
                "representative_combination_id": combination_id,
                "representative_decision_path": " | ".join(result.decision_path),
                "example_inputs": input_values,
            },
        )
        recommendation_set["valid_combination_count"] += 1
        recommendation_set["primary_tasks"].add(inputs.primary_task)
        recommendation_set["data_forms"].add(inputs.data_form)
        recommendation_set["target_audiences"].add(inputs.target_audience)
        recommendation_set["display_levels"].add(inputs.display_level)
        recommendation_set["comparison_foci"].add(inputs.comparison_focus)
        recommendation_set["comparison_structures"].add(inputs.comparison_structure)
        recommendation_set["temporal_contexts"].add(inputs.temporal_context)
        recommendation_set["n_overlaid_series_values"].add(inputs.n_overlaid_series)
        recommendation_set["n_comparison_levels_values"].add(inputs.n_comparison_levels)
        recommendation_set["n_compositional_parts_values"].add(
            inputs.n_compositional_parts
        )

        valid_rows.append(
            {
                "combination_id": combination_id,
                **input_values,
                "decision_path": " | ".join(result.decision_path),
                "recommendation_count": len(result.recommendations),
                "recommendation_names": " | ".join(names),
                "recommendation_ranks": " | ".join(ranks),
                "recommendation_output": format_result(result),
                "design_notes": notes,
            }
        )

        task_counts[inputs.primary_task] += 1
        for number, rec in enumerate(result.recommendations, start=1):
            recommendation_counts[rec.visualisation] += 1
            long_rows.append(
                {
                    "combination_id": combination_id,
                    **input_values,
                    "recommendation_number": number,
                    **asdict(rec),
                    "design_notes": notes,
                }
            )

    validation_rows = list(rejected_examples.values())
    validation_headers = [
        "example",
        *INPUT_FIELDS,
        "corrective_error_message",
    ]
    set_rows = recommendation_set_rows(recommendation_sets)

    write_csv(VALID_CSV, VALID_HEADERS, valid_rows)
    write_csv(LONG_CSV, LONG_HEADERS, long_rows)
    write_csv(RECOMMENDATION_SET_CSV, RECOMMENDATION_SET_HEADERS, set_rows)
    write_csv(VALIDATION_CSV, validation_headers, validation_rows)

    summary = {
        "candidate_combinations_evaluated": candidate_count,
        "valid_decision_equivalent_combinations": len(valid_rows),
        "rejected_candidate_combinations": candidate_count - len(valid_rows),
        "long_format_recommendation_rows": len(long_rows),
        "unique_visualisation_names": len(recommendation_counts),
        "unique_ordered_recommendation_name_sets": len(
            unique_ordered_recommendation_name_sets
        ),
        "unique_reviewable_recommendation_sets": len(recommendation_sets),
        "unique_recommendation_detail_records": len(
            unique_recommendation_detail_records
        ),
        "unique_recommendation_and_guidance_outputs": len(
            unique_recommendation_and_guidance_outputs
        ),
        "unique_complete_formatted_outputs": len(unique_complete_formatted_outputs),
        "unique_decision_paths": len(unique_decision_paths),
        "unique_design_note_sets": len(unique_design_note_sets),
        "valid_combinations_by_primary_task": dict(sorted(task_counts.items())),
        "recommendation_frequency": dict(
            sorted(recommendation_counts.items(), key=lambda item: (-item[1], item[0]))
        ),
        "rejection_reason_frequency": dict(
            sorted(rejected_reasons.items(), key=lambda item: (-item[1], item[0]))
        ),
        "enumeration_domains": enumeration_domains(),
        "representative_numeric_values": REPORT_REPRESENTATIVE_VALUES,
    }
    workbook_data = {
        "summary": summary,
        "summary_metric_rows": summary_metric_rows(summary),
        "task_rows": [
            {"primary_task": task, "valid_combinations": count}
            for task, count in sorted(task_counts.items())
        ],
        "recommendation_frequency_rows": [
            {"recommendation_name": name, "frequency": count}
            for name, count in sorted(
                recommendation_counts.items(), key=lambda item: (-item[1], item[0])
            )
        ],
        "how_to_use_rows": how_to_use_rows(),
        "rejected_reason_rows": [
            {"reason": reason, "count": count}
            for reason, count in sorted(
                rejected_reasons.items(), key=lambda item: (-item[1], item[0])
            )
        ],
        "valid_headers": VALID_HEADERS,
        "valid_rows": valid_rows,
        "long_headers": LONG_HEADERS,
        "long_rows": long_rows,
        "recommendation_set_headers": RECOMMENDATION_SET_HEADERS,
        "recommendation_set_rows": set_rows,
        "validation_headers": validation_headers,
        "validation_rows": validation_rows,
        "parameter_rows": parameter_report_rows(),
    }
    SUMMARY_JSON.write_text(json.dumps(summary, indent=2), encoding="utf-8")
    WORKBOOK_JSON.write_text(json.dumps(workbook_data), encoding="utf-8")
    write_excel_workbook(workbook_data)
    write_markdown_report(workbook_data)

    print(json.dumps(summary, indent=2))


if __name__ == "__main__":
    main()
